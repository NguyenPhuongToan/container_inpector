from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.core.constants import REPORT_ROOT
from app.database.models import Inspection


class ExcelExportError(RuntimeError):
    """Raised when an Excel export cannot be generated."""


def _inspection_to_export_dict(inspection: Inspection) -> dict[str, Any]:
    images = [
        {
            "angle": image.angle,
            "label": image.label,
            "url": image.url,
            "path": image.path,
        }
        for image in sorted(inspection.images, key=lambda item: item.angle)
    ]
    return {
        "container_number": inspection.container_number,
        "booking_number": inspection.booking_number,
        "truck_number": inspection.truck_number,
        "worker_name": inspection.worker_name,
        "port_name": inspection.port_name,
        "notes": inspection.notes or "",
        "status": inspection.status,
        "created_at": inspection.created_at.isoformat(),
        "images": images,
    }


def _report_dir(inspection: Inspection) -> Path:
    report_dir = REPORT_ROOT / inspection.id
    report_dir.mkdir(parents=True, exist_ok=True)
    return report_dir


def generate_excel(inspection: Inspection) -> Path:
    try:
        inspection_data = _inspection_to_export_dict(inspection)
        report_path = _report_dir(inspection) / f"inspection_{inspection.id}_excel.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Inspection"
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 44
        sheet.column_dimensions["C"].width = 16
        sheet.column_dimensions["D"].width = 26

        title_fill = PatternFill("solid", fgColor="075DCC")
        sheet["A1"] = "Container Inspection Report"
        sheet["A1"].font = Font(bold=True, color="FFFFFF", size=16)
        sheet["A1"].fill = title_fill
        sheet.merge_cells("A1:D1")

        rows = [
            ("Container Number", inspection_data["container_number"]),
            ("Booking Number", inspection_data["booking_number"]),
            ("Truck Number", inspection_data["truck_number"]),
            ("Worker", inspection_data["worker_name"]),
            ("Port / Location", inspection_data["port_name"]),
            ("Status", inspection_data["status"]),
            ("Submitted At", inspection_data["created_at"]),
            ("Notes", inspection_data.get("notes", "")),
        ]

        for row_index, (label, value) in enumerate(rows, start=3):
            sheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
            sheet.cell(row=row_index, column=2, value=value)

        start_row = 13
        headers = ["Angle", "Label", "Image URL", "Saved File"]
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=start_row, column=column, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = title_fill

        for offset, image in enumerate(inspection_data.get("images", []), start=1):
            row = start_row + offset
            sheet.cell(row=row, column=1, value=image["angle"])
            sheet.cell(row=row, column=2, value=image["label"])
            sheet.cell(row=row, column=3, value=image["url"])
            sheet.cell(row=row, column=4, value=image.get("path", ""))

        workbook.save(report_path)
        return report_path
    except Exception as exc:
        raise ExcelExportError("Could not generate Excel export") from exc
